from SuperSort import *
import random
import openpyxl


def create_weights(all_unique_words):
    random_weight = []
    for i in range(all_unique_words):
        kek = random.randint(1, 9)
        random_weight.append(kek)

    return random_weight


# количество всех отзывов + кол-во уникальных слов для использования в циклах
all_reviews_len = len(all_revs)
all_unique_words = len(dataset_words)
print(f"Всего отзывов: {all_reviews_len} \nВсего уникальных слов: {all_unique_words}")
workbook = openpyxl.load_workbook("результат.xlsx")
binaryTable = workbook.active

# Считываем данные из ячеек A3 до A502
reviews = []
for row in binaryTable.iter_rows(min_row=3, max_row=all_reviews_len + 2, min_col=0, max_col=0, values_only=True):
    reviews.append(row[0])


# Считываем значения из ячеек B3 до HH3 и записываем в массив
binaryReview = []
for row in binaryTable.iter_rows(min_row=3, max_row=all_reviews_len + 2, min_col=2, max_col=all_unique_words + 1, values_only=True):
    binaryReview.append(list(row))

EPOCH = 1
sumBorder = 1000
count = 0
shuffled_indices = list(range(len(reviews)))
true_rev = []
false_rev = []
all_rev = ['0T', '1F', '2T', '3F', '4T', '5F', '6T', '7F', '8T', '9F', '10T', '1F', '12T', '13F', '14T', '15F', '16T', '17F', '18T', '19F', '20T', '21F', '22T', '23F', '24T', '25F', '26T', '27F', '28T', '29F', '30T', '31F', '32T', '33F', '34T', '35F', '36T', '37F', '38T', '39F', '40T', '41F', '42T', '43F', '44T', '45F', '46T', '47F', '48T', '49F', '50T', '51F', '52T', '53F', '54T', '55F', '56T', '57F', '58T', '59F', '60T', '61F', '62T', '63F', '64T', '65F', '66T', '67F', '68T', '69F', '70T', '71F', '72T', '73F', '74T', '75F', '76T', '77F', '78T', '79F', '80T', '81F', '82T', '83F', '84T', '85F', '86T', '87F', '88T', '89F', '90T', '91F', '92T', '93F', '94T', '95F', '96T', '97F', '98T', '99F', '100T', '101F', '102T', '103F', '104T', '105F', '106T', '107F', '108T', '109F', '110T', '111F', '112T', '113F', '114T', '115F', '116T', '117F', '118T', '119F', '120T', '121F', '122T', '123F', '124T', '125F', '126T', '127F', '128T', '129F', '130T', '131F', '132T', '133F', '134T', '135F', '136T', '137F', '138T', '139F', '140T', '141F', '142T', '143F', '144T', '145F', '146T', '147F', '148T', '149F', '150T', '151F', '152T', '153F', '154T', '155F', '156T', '157F', '158T', '159F', '160T', '161F', '162T', '163F', '164T', '165F', '166T', '167F', '168T', '169F', '170T', '171F', '172T', '173F', '174T', '175F', '176T', '177F', '178T', '179F', '180T', '181F', '182T', '183F', '184T', '185F', '186T', '187F', '188T', '189F', '190T', '191F', '192T', '193F', '194T', '195F', '196T', '197F', '198T', '199F', '200T', '201F', '202T', '203F', '204T', '205F', '206T', '207F', '208T', '209F', '210T', '211F', '212T', '213F', '214T', '215F', '216T', '217F', '218T', '219F', '220T', '221F', '222T', '223F', '224T', '225F', '226T', '227F', '228T', '229F', '230T', '231F', '232T', '233F', '234T', '235F', '236T', '237F', '238T', '239F', '240T', '241F', '242T', '243F', '244T', '245F', '246T', '247F', '248T', '249F', '250T', '251F', '252T', '253F', '254T', '255F', '256T', '257F', '258T', '259F', '260T', '261F', '262T', '263F', '264T', '265F', '266T', '267F', '268T', '269F', '270T', '271F', '272T', '273F', '274T', '275F', '276T', '277F', '278T', '279F', '280T', '281F', '282T', '283F', '284T', '285F', '286T', '287F', '288T', '289F', '290T', '291F', '292T', '293F', '294T', '295F', '296T', '297F', '298T', '299F', '300T', '301F', '302T', '303F', '304T', '305F', '306T', '307F', '308T', '309F', '310T', '311F', '312T', '313F', '314T', '315F', '316T', '317F', '318T', '319F', '320T', '321F', '322T', '323F', '324T', '325F', '326T', '327F', '328T', '329F', '330T', '331F', '332T', '333F', '334T', '335F', '336T', '337F', '338T', '339F', '340T', '341F', '342T', '343F', '344T', '345F', '346T', '347F', '348T', '349F', '350T', '351F', '352T', '353F', '354T', '355F', '356T', '357F', '358T', '359F', '360T', '361F', '362T', '363F', '364T', '365F', '366T', '367F', '368T', '369F', '370T', '371F', '372T', '373F', '374T', '375F', '376T', '377F', '378T', '379F', '380T', '381F', '382T', '383F', '384T', '385F', '386T', '387F', '388T', '389F', '390T', '391F', '392T', '393F', '394T', '395F', '396T', '397F', '398T', '399F', '400T', '401F', '402T', '403F', '404T', '405F', '406T', '407F', '408T', '409F', '410T', '411F', '412T', '413F', '414T', '415F', '416T', '417F', '418T', '419F', '420T', '421F', '422T', '423F', '424T', '425F', '426T', '427F', '428T', '429F', '430T', '431F', '432T', '433F', '434T', '435F', '436T', '437F', '438T', '439F', '440T', '441F', '442T', '443F', '444T', '445F', '446T', '447F', '448T', '449F', '450T', '451F', '452T', '453F', '454T', '455F', '456T', '457F', '458T', '459F', '460T', '461F', '462T', '463F', '464T', '465F', '466T', '467F', '468T', '469F', '470T', '471F', '472T', '473F', '474T', '475F', '476T', '477F', '478T', '479F', '480T', '481F', '482T', '483F', '484T', '485F', '486T', '487F', '488T', '489F', '490T', '491F', '492T', '493F', '494T', '495F', '496T', '497F', '498T', '499F', '500T', '501F', '502T', '503F', '504T', '505F', '506T', '507F', '508T', '509F', '510T', '511F', '512T', '513F', '514T', '515F', '516T', '517F', '518T', '519F', '520T', '521F', '522T', '523F', '524T', '525F', '526T', '527F', '528T', '529F', '530T', '531F', '532T', '533F', '534T', '535F', '536T', '537F', '538T', '539F', '540T', '541F', '542T', '543F', '544T', '545F', '546T', '547F', '548T', '549F', '550T', '551F', '552T', '553F', '554T', '555F', '556T', '557F', '558T', '559F', '560T', '561F', '562T', '563F', '564T', '565F', '566T', '567F', '568T', '569F', '570T', '571F', '572T', '573F', '574T', '575F', '576T', '577F', '578T', '579F', '580T', '581F', '582T', '583F', '584T', '585F', '586T', '587F', '588T', '589F', '590T', '591F', '592T', '593F', '594T', '595F', '596T', '597F', '598T', '599F', '600T', '601F', '602T', '603F', '604T', '605F', '606T', '607F', '608T', '609F', '610T', '611F', '612T', '613F', '614T', '615F', '616T', '617F', '618T', '619F', '620T', '621F', '622T', '623F', '624T', '625F', '626T', '627F', '628T', '629F', '630T', '631F', '632T', '633F', '634T', '635F', '636T', '637F', '638T', '639F', '640T', '641F', '642T', '643F', '644T', '645F', '646T', '647F', '648T', '649F', '650T', '651F', '652T', '653F', '654T', '655F', '656T', '657F', '658T', '659F', '660T', '661F', '662T', '663F', '664T', '665F', '666T', '667F', '668T', '669F', '670T', '671F', '672T', '673F', '674T', '675F', '676T', '677F', '678T', '679F', '680T', '681F', '682T', '683F', '684T', '685F', '686T', '687F', '688T', '689F', '690T', '691F', '692T', '693F', '694T', '695F', '696T', '697F', '698T', '699F']#all_rev = [str(i) + ("T" if i < len(positive_reviews) else "F") for i in shuffled_indices]
labels = []
# #j = 350
for i in range(700):
   if i % 2 == 0:
       labels.append(1)
   else:
       labels.append(0)

print(labels)

max_correct_ans = -100
perfect_weights = []
nicePercent = 0
filename = 'result.xlsx'
# Создайте или откройте файл Excel
workbook = openpyxl.Workbook()
sheet = workbook.active
